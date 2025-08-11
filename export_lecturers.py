# Exportiert alle Dozenten in eine Excel-Datei
import logging

from openpyxl import Workbook
from config import Config
from studi_manager_controller import StudentManagerController
from studi_manager_model import Model

myModel = Model()
myController = StudentManagerController(myModel)

# Konfigurationsdaten holen
configs = Config()
# Logger initialisieren
logger = logging.getLogger(__name__)
logging.basicConfig(
    filename=configs.log_file_path + configs.log_file_name,
    encoding="utf-8",
    level=configs.log_level,
    format=configs.log_format,
)
logger.info("Export gestartet")

# Excel File aufbauen
logger.debug("Create Workbook")
wb = Workbook()
ws = wb.active
ws.title = "Dozenten Export"
titles = "Nachname Dozent, Vorname Dozent, Email Dozent, Dozenten-ID, Personen-ID, Firma, Ist Gutachter".split(
    ", "
)
for col, title in enumerate(titles, start=1):
    ws.cell(row=1, column=col).value = title

logger.info("Lese alle Gutachter ein")
lecturers = myController.read_all_lecturers()

row_idx = 1
for lecturer in lecturers:
    logger.debug("Schreibe Gutachter: %s %s", lecturer.first_name, lecturer.last_name)
    ws.cell(row=row_idx, column=1).value = lecturer.last_name
    ws.cell(row=row_idx, column=2).value = lecturer.first_name
    ws.cell(row=row_idx, column=3).value = lecturer.email
    ws.cell(row=row_idx, column=4).value = lecturer.lecturer_id
    ws.cell(row=row_idx, column=5).value = lecturer.person_id
    ws.cell(row=row_idx, column=6).value = lecturer.company
    ws.cell(row=row_idx, column=7).value = lecturer.is_reviewer
    # Row Index erhöhen
    row_idx += 1

logger.debug("Ausgabedatei schreiben")
FILE_NAME = "Dozenten.xlsx"
SAVE_PATH = "."
save_file_path = SAVE_PATH + "/" + FILE_NAME
wb.save(save_file_path)
logger.info("Es wurden %s Gutacher exportiert", row_idx - 1)
