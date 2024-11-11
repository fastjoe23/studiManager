# Kleines Skript um alle Assignements eines bestimmten Kurses zu exportieren

from openpyxl import Workbook
from studi_manager_controller import StudentManagerController
from studi_manager_model import Model
from dhbw_mail import DHBWMail

# Variablen festlegen
course_id = 2
assignment_type = "Projektarbeit 2"
save_path = "."

myModel = Model()
myController = StudentManagerController(myModel)

mail_client = DHBWMail()

#Excel File aufbauen
wb = Workbook()
ws = wb.active
# Kurs einlesen
course = myController.read_course_by_id(course_id)
ws.title = course.course_name
titles = "Typ, Nachname Student, Vorname Student, Thema, Nachname Gutachter, Vorname Gutachter, Email Student, Email Gutachter, Email-Text".split(
            ", "
        )
for col, title in enumerate(titles, start=1):
    ws.cell(row=1, column=col).value = title

# alle Studenten aus dem Kurs lesen
students_list = myController.read_all_students_by_course_id(course_id)

# Schleife ueber alle Studenten im Kurs
row_idx = 2
for student in students_list:
    if student.enrolled:
        # Assignment ermitteln
        assignment = myController.read_assignment_by_student_id_and_type(
            student.student_id, assignment_type
        )
        if assignment:
            # Gutachter ermitteln
            lecturer = myController.read_lecturer_by_id(assignment.lecturer_id)
            if not lecturer:
                print(
                    f"Gutachter {assignment.lecturer_id} nicht gefunden."
                )
            # Platzhalter ersetzen
            # Variablen für die Ersetzung
            variables = {
                "Vorname_student": student.first_name,
                "Nachname_student": student.last_name,
                "Vorname_gutachter": lecturer.first_name,
                "Nachname_gutachter": lecturer.last_name,
            }
            # Ersetze die Platzhalter in der Zeichenkette durch die Variablen
            email_text = myController.get_mail_text(assignment_type)
            formatted_email_text = email_text.format(**variables)

            # Schreibe in Ausgabefile
            ws.cell(row=row_idx, column=1).value = assignment_type
            ws.cell(row=row_idx, column=2).value = student.last_name
            ws.cell(row=row_idx, column=3).value = student.first_name
            ws.cell(row=row_idx, column=4).value = assignment.topic
            ws.cell(row=row_idx, column=5).value = lecturer.last_name
            ws.cell(row=row_idx, column=6).value = lecturer.first_name
            ws.cell(row=row_idx, column=7).value = student.email
            ws.cell(row=row_idx, column=8).value = lecturer.email
            ws.cell(row=row_idx, column=9).value = formatted_email_text

            # Row Index erhöhen
            row_idx += 1

file_name = course.course_name + assignment_type + ".xlsx"
save_file_path = save_path + "/" + file_name
wb.save(save_file_path)

            


