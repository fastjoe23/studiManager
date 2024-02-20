# Kleines Skript um Gutachter aus einer Excel zu importieren
from openpyxl import load_workbook

from studi_manager_controller import StudentManagerController
from studi_manager_model import Model

FILE_PATH = ""
workbook = load_workbook(FILE_PATH)

ws = workbook.active
model = Model()
controller = StudentManagerController(model)
imported_lecturers = []
for row in ws.iter_rows(min_row=2, values_only=True):
    firstname = row[0]
    lastname = row[1]
    email = row[2]
    company = row[3] if row[3] else ""
    print(f"Importiere {firstname} {lastname} {email} {company}")
    try:
        lecturer = controller.add_lecturer(lastname, firstname, email, company)
        imported_lecturers.append(lecturer)
    except Exception as e:
        raise e
    
print(f"Es wurden {len(imported_lecturers)} erfolgreich importiert.")
