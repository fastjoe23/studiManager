# Erzeugt aus einer Eingabeliste Bewertungsbögen
import logging

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from config import Config
from presentation_evaluation_word import PresentationEvaluationDocx

# Input-Datei festlegen
INPUT_FILE = "./tests/create_evaluation_projekt2025.xlsx"
SAVE_PATH = "C:/Users/Offtermatt/Downloads/EvaluationSheets"
COURSE_NAME = "WWI2022V"
SIGNATURE = False  # Unterschrift hinzufügen oder nicht



# Konfigurationsdaten holen
configs = Config()
# Logger initialisieren
logger = logging.getLogger(__name__)
logging.basicConfig(
    filename=configs.log_file_path + configs.log_file_name,
    encoding="utf-8",
    level=configs.log_level,
    format=configs.log_format,
)
logger.info("Erzeugung Bewertungsbögen gestartet")
try:
    workbook = load_workbook(INPUT_FILE)
    logger.info("Eingabe-Datei erfolgreich gelesen:%s",  INPUT_FILE)
except InvalidFileException:
    logger.error("Eingabedatei nicht gefunden: %s", INPUT_FILE)
    raise

ws = workbook.active

for row in ws.iter_rows(min_row=2, values_only=True):
    assignment_type = row[0]
    name_of_students = row[1]
    topic = row[2]
    assignment_date = row[3]
    assignment_time = row[4]
    logger.info("Erzeuge Word für %s mit Thema %s", name_of_students, topic)
    word_file = PresentationEvaluationDocx(
    SAVE_PATH,
    assignment_type,
    COURSE_NAME,
    topic,
    [],
    name_of_students,
    assignment_date.strftime("%d.%m.%Y"),
    assignment_time,
    SIGNATURE 
    )
    try:
        pdf_file_path = word_file.create_evaluation()
        logger.info(
            "Evaluations-PDF mit Thema %sfür Studente %s erzeugt.",
            topic,
            name_of_students,
        )
    except Exception as e:
        logger.error("Error occured: %s", e)
        raise e