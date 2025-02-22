import csv
from datetime import datetime
import re
import logging
from openpyxl import Workbook, load_workbook
from presentation_evaluation_pdf import PresentationEvaluationPDF
from studi_manager_model import LastUsedItems, Person
from dhbw_mail import DHBWMail


# Hilfsklassen
class StudentNotFoundException(Exception):
    pass


class LecturerNotFoundException(Exception):
    pass


class StudentManagerController:
    def __init__(self, model):
        self.model = model
        self.logger = logging.getLogger("main")
        self.logger.debug("Controller called")

    # Controller Methoden fuer Person
    def add_person(self, last_name, first_name, email):
        new_person = Person()
        if email:
            self.check_valid_mail(email)
        return new_person.create_person(last_name, first_name, email)

    def delete_person(self, person_id):

        return self.model.person.delete_person(person_id)

    def update_person(self, person_id, last_name, first_name, email):
        if email:
            self.check_valid_mail(email)
        return self.model.person.update_person(person_id, last_name, first_name, email)

    def read_person_by_id(self, person_id):
        return self.model.person.read_person_by_id(person_id)

    def read_person_by_name(self, last_name, first_name):
        return self.model.person.read_person_by_name(last_name, first_name)

    def read_all_persons(self):
        return self.model.person.read_all_persons()

    # Controller Methoden fuer Student
    def add_student(self, last_name, first_name, email, company, mat_number, enrolled):
        if email:
            self.check_valid_mail(email)

        new_student = self.model.student.create_student(
            last_name, first_name, email, company, mat_number, enrolled
        )
        return new_student

    def delete_student(self, student_id):
        return self.model.student.delete_student(student_id)

    def update_student(
        self,
        student_id,
        person_id,
        last_name,
        first_name,
        email,
        company,
        mat_number,
        enrolled,
    ):
        if email:
            self.check_valid_mail(email)
        return self.model.student.update_student(
            student_id,
            person_id,
            last_name,
            first_name,
            email,
            company,
            mat_number,
            enrolled,
        )

    def read_student_by_id(self, student_id):
        return self.model.student.read_student_by_id(student_id)

    def read_student_by_person_id(self, person_id):
        return self.model.student.read_student_by_person_id(person_id)

    def read_student_by_name(self, last_name, first_name):
        return self.model.student.read_student_by_name(last_name, first_name)

    def read_all_students(self):
        return self.model.student.read_all_students()

    def read_all_students_by_course_id(self, course_id):
        return self.model.student.read_all_students_by_course_id(course_id)

    # Methoden für Lecturer
    def add_lecturer(self, last_name, first_name, email, company):
        if email:
            self.check_valid_mail(email)

        new_lecturer = self.model.lecturer.create_lecturer(
            last_name, first_name, email, company
        )
        return new_lecturer

    def delete_lecturer(self, lecturer_id):
        return self.model.lecturer.delete_lecturer(lecturer_id)

    def update_lecturer(
        self, lecturer_id, person_id, last_name, first_name, email, company
    ):
        return self.model.lecturer.update_lecturer(
            lecturer_id, person_id, last_name, first_name, email, company
        )

    def read_lecturer_by_id(self, lecturer_id):
        return self.model.lecturer.read_lecturer_by_id(lecturer_id)

    def read_lecturer_by_person_id(self, person_id):
        return self.model.lecturer.read_lecturer_by_person_id(person_id)

    def read_lecturer_by_name(self, last_name, first_name):
        return self.model.lecturer.read_lecturer_by_name(last_name, first_name)

    def read_all_lecturers(self):
        return self.model.lecturer.read_all_lecturers()

    def check_valid_mail(self, email):
        # Einfache Validierung für eine gültige E-Mail-Adresse
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            self.logger.warning("Ungültige Email-Adresse: %s", email)
            raise ValueError("Ungültige E-Mail-Adresse")

    # Methoden für Course
    def add_course(self, course_name, start_date):
        new_course = self.model.course.create_course(course_name, start_date)
        return new_course

    def delete_course(self, course_id):
        return self.model.course.delete_course(course_id)

    def update_course(self, course_id, course_name, start_date):
        return self.model.course.update_course(course_id, course_name, start_date)

    def read_course_by_id(self, course_id):
        return self.model.course.read_course_by_id(course_id)

    def read_all_courses(self):
        return self.model.course.read_all_courses()

    # Methoden für Enrollments
    def add_student_to_course(self, student_id, course_id):
        return self.model.enrollments.add_student_to_course(student_id, course_id)

    def read_all_enrollments(self):
        return self.model.enrollments.read_all_enrollments()

    def read_all_enrollments_by_student_id(self, student_id):
        return self.model.enrollments.read_all_enrollments_by_student_id(student_id)

    def delete_enrollment(self, enrollment_id):
        return self.model.enrollments.delete_enrollment(enrollment_id)

    def update_enrollment(self, enrollment_id, student_id, course_id):
        return self.model.enrollments.update_enrollment(
            enrollment_id, student_id, course_id
        )

    def read_enrollment_by_id(self, enrollment_id):
        return self.model.enrollments.read_enrollment_by_id(enrollment_id)

    # Methoden für Assigments
    def create_assignment(
        self, student_id, lecturer_id, assignment_type, topic, grade, assignment_date, assignment_time
    ):
        return self.model.assignments.create_assignment(
            student_id, lecturer_id, assignment_type, topic, grade, assignment_date, assignment_time
        )

    def add_student_to_assignment(self, student_id, assignment_id):
        return self.model.assignments.add_student_to_assignment(
            student_id, assignment_id
        )

    def add_lecturer_to_assignment(self, lecturer_id, assignment_id):
        return self.model.assignments.add_student_to_assignment(
            lecturer_id, assignment_id
        )

    def read_all_assignments(self):
        return self.model.assignments.read_all_assignments()

    def read_all_assignments_by_student_id(self, student_id):
        return self.model.assignments.read_all_assignments_by_student_id(student_id)

    def read_all_assignments_by_lecturer_id(self, lecturer_id):
        return self.model.assignments.read_all_assignments_by_lecturer_id(lecturer_id)

    def delete_assignment(self, assignment_id):
        return self.model.assignments.delete_assignment(assignment_id)

    def update_assignment(
        self,
        assignment_id,
        student_id,
        lecturer_id,
        assignment_type,
        topic,
        grade,
        assignment_date,
        assignment_time,
    ):
        return self.model.assignments.update_assignment(
            assignment_id,
            student_id,
            lecturer_id,
            assignment_type,
            topic,
            grade,
            assignment_date,
            assignment_time,
        )

    def read_assignment_by_id(self, assignment_id):
        return self.model.assignments.read_assignment_by_id(assignment_id)

    def read_assignment_by_student_id_and_type(self, student_id, assignment_type):
        return self.model.assignments.read_assignment_by_student_id_and_type(
            student_id, assignment_type
        )

    # Methoden für last_used_items
    def create_last_used_item(self, item_type, elements):
        return self.model.last_used_items.create_last_used_item(item_type, elements)

    def read_last_used_item_by_type(self, item_type):
        return self.model.last_used_items.read_last_used_item_by_type(item_type)

    def update_last_used_item(self, item_type, elements):
        return self.model.last_used_items.update_last_used_item(item_type, elements)

    def add_element_to_last_used_items(self, item_type, element):
        actual_state = LastUsedItems()
        actual_state = self.read_last_used_item_by_type(item_type)

        if actual_state:  # es gibt schon einen Eintrag von diesem Type
            if element in actual_state.elements:
                # das element ist schon in den zuletzt verwendeten Elementen wir setzen es also nur ans Ende der Liste
                # als zuletzt zuletzt verwendetes Element
                # erstmal löschen
                actual_state.elements.remove(element)
                # jetzt ans Ende setzen
                actual_state.elements.append(element)
            else:
                # das Element ist noch nicht in der Liste und die Liste ist noch nicht voll
                if len(actual_state.elements) < 3:
                    actual_state.elements.append(element)
                else:
                    # Es sind schon drei Elemente drin also löschen wir das erste Element (es ist das älteste)
                    # und setzen das neue ans Ende
                    actual_state.elements.pop(0)
                    actual_state.elements.append(element)

            # jetzt aktualisieren wir den Eintrag in der Datenbank
            self.update_last_used_item(actual_state.type, actual_state.elements)
        else:
            # es gibt noch keinen Eintrag zu diesem Type also erstellen wir einen
            new_elements = [element]
            self.create_last_used_item(item_type, new_elements)

    # Methoden für Notizen
    def create_note(self, new_note):
        return self.model.note.create_note(new_note)

    def read_all_notes(self):
        return self.model.note.read_all_notes()

    def read_notes_by_type_and_related_id(self, note_type, related_id):
        return self.model.note.read_notes_by_type_and_related_id(note_type, related_id)

    def update_note_by_id(self, note_id, new_note):
        return self.model.note.update_note_by_id(note_id, new_note)

    def read_note_by_id(self, note_id):
        return self.model.note.read_note_by_id(note_id)

    def delete_note_by_id(self, note_id):
        return self.model.note.delete_note_by_id(note_id)

    def delete_all_notes(self):
        return self.model.note.delete_all_notes()

    # Import Methoden
    def save_list_to_file(self, parsed_data, save_path, assignment_type, course_name):
        file_name = course_name + assignment_type + ".csv"
        # Speichern des geparsten Ergebnisses in dem ausgewählten Ordner
        try:
            save_file_path = save_path + "/" + file_name
            with open(save_file_path, "w", encoding="utf-8") as file:
                file.write(parsed_data)
        except Exception as e:
            self.logger.error("Error occurred: %s", e)
            raise e

    def save_excel_to_file(self, workbook, save_path, assignment_type, course_name):
        file_name = course_name + assignment_type + ".xlsx"
        # Speichern des geparsten Ergebnisses in dem ausgewählten Ordner
        try:
            save_file_path = save_path + "/" + file_name
            workbook.save(save_file_path)
        except Exception as e:
            self.logger.error("Error occurred: %s", e)
            raise e

    def parse_to_csv(self, text_content):
        # Hier die eigentliche Parsing-Logik implementieren
        # Verwendet Regular Expressions, um Informationen aus den Zeilen zu extrahieren
        pattern = r"(\w+)\s+\(([^,]+),\s*([^)]+)\)"
        matches = re.findall(pattern, text_content)

        # Extrahiere Informationen und erstelle einen formatierten Text
        parsed_result = "Nachname,Vorname,Firma,Mat-Nr,Email\n"
        for match in matches:
            email, lastname, firstname = match
            parsed_result += (
                f"{lastname},{firstname},,,{email}@lehre.dhbw-stuttgart.de\n"
            )

        return parsed_result

    def parse_to_excel(self, text_content):
        # Hier die eigentliche Parsing-Logik implementieren
        # Verwendet Regular Expressions, um Informationen aus den Zeilen zu extrahieren
        pattern = r"(\w+)\s+\(([^,]+),\s*([^)]+)\)"
        matches = re.findall(pattern, text_content)

        # Erstelle ein neues Workbook
        workbook = Workbook()
        sheet = workbook.active

        # Schreibe Titelzeile
        sheet["A1"] = "Nachname"
        sheet["B1"] = "Vorname"
        sheet["C1"] = "Firma"
        sheet["D1"] = "Mat-Nr"
        sheet["E1"] = "Email"

        # Schreibe Daten in die Spalten
        row_idx = 0
        for row_idx, match in enumerate(matches, start=2):
            email, lastname, firstname = match
            sheet.cell(row=row_idx, column=1).value = lastname
            sheet.cell(row=row_idx, column=2).value = firstname
            # Firma und Mat-Nr leer lassen
            sheet.cell(row=row_idx, column=3).value = ""
            sheet.cell(row=row_idx, column=4).value = ""
            sheet.cell(row=row_idx, column=5).value = f"{email}@lehre.dhbw-stuttgart.de"
            self.logger.debug("Student %s %s %s geparst.", firstname, lastname, email)

        self.logger.info("Es wurden %s Zeilen erfolgreich geparst", row_idx - 1)

        return workbook

    def import_students_from_csv_into_course(self, csv_path, course_id):
        try:
            with open(csv_path, "r", encoding="utf-8") as file:
                csv_reader = csv.reader(file)
                next(csv_reader)  # Überspringe Header, wenn vorhanden

                for row in csv_reader:
                    lastname, firstname, company, mat_number, email = [
                        value.strip() for value in row
                    ]
                    new_student = self.add_student(
                        lastname, firstname, email, company, mat_number, True
                    )
                    self.add_student_to_course(new_student.student_id, course_id)

        except Exception as e:
            self.logger.error("Error occurred: %s", e)
            raise e

    def import_students_from_excel_into_course(self, excel_path, course_id):
        try:
            workbook = load_workbook(excel_path)
            sheet = workbook.active

            counter = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                lastname = row[0]
                firstname = row[1]
                company = row[2]
                mat_number = row[3]
                email = row[4]
                new_student = self.add_student(
                    lastname, firstname, email, company, mat_number, True
                )
                self.logger.debug("Student %s %s geparst.", firstname, lastname)
                self.add_student_to_course(new_student.student_id, course_id)
                counter += 1

            workbook.close()
            self.logger.info("Es wurden %s Studenten importiert", counter)

        except Exception as e:
            self.logger.error("Error occurred: %s", e)
            raise e

    def import_assignments_from_csv_into_course(self, csv_path):
        try:
            with open(csv_path, "r", encoding="utf-8") as file:
                csv_reader = csv.reader(file)
                next(csv_reader)  # Überspringe Header, wenn vorhanden

                imported_assignments = []

                for row in csv_reader:
                    (
                        assignment_type,
                        last_name_student,
                        first_name_student,
                        topic,
                        last_name_lecturer,
                        first_name_lecturer,
                        grade,
                        assignment_date,
                        assignment_time,
                    ) = [value.strip() for value in row]
                    # Studentendaten holen
                    student = self.read_student_by_name(
                        last_name_student, first_name_student
                    )
                    if not student:
                        self.logger.error(
                            "Student: %s %s nicht gefunden",
                            first_name_student,
                            last_name_student,
                        )
                        raise StudentNotFoundException(
                            f"Student {first_name_student} {last_name_student} nicht gefunden."
                        )
                    # Gutachterdaten holen
                    lecturer = self.read_lecturer_by_name(
                        last_name_lecturer, first_name_lecturer
                    )
                    if not lecturer:
                        self.logger.error(
                            "Dozent: %s %s nicht gefunden",
                            first_name_lecturer,
                            last_name_lecturer,
                        )
                        raise LecturerNotFoundException(
                            f"Gutachter {first_name_lecturer} {last_name_lecturer} nicht gefunden."
                        )

                    # jetzt alle Daten zusammen also Assignment erstellen
                    assignment = self.create_assignment(
                        student.student_id,
                        lecturer.lecturer_id,
                        assignment_type,
                        topic,
                        grade,
                        assignment_date,
                        assignment_time,
                    )
                    imported_assignments.append(assignment)
                    self.logger.debug(
                        "Assignment %s für Student Id %s importiert.",
                        assignment_type,
                        student.student_id,
                    )

            return imported_assignments
        except Exception as e:
            self.logger.error("Error occurred: %s", e)
            raise e

    def import_assignments_from_excel_into_course(self, excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active

            imported_assignments = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                assignment_type = row[0]
                last_name_student = row[1]
                first_name_student = row[2]
                topic = row[3]
                last_name_lecturer = row[4]
                first_name_lecturer = row[5]
                grade = row[6]
                assignment_date = row[7]
                assignment_time = row[8]

                # Studentendaten holen
                student = self.read_student_by_name(
                    last_name_student, first_name_student
                )
                if not student:
                    self.logger.error(
                        "Student: %s %s nicht gefunden",
                        first_name_student,
                        last_name_student,
                    )
                    raise StudentNotFoundException(
                        f"Student {first_name_student} {last_name_student} nicht gefunden."
                    )
                # Gutachterdaten holen
                lecturer = self.read_lecturer_by_name(
                    last_name_lecturer, first_name_lecturer
                )
                if not lecturer:
                    self.logger.error(
                        "Dozent: %s %s nicht gefunden",
                        first_name_lecturer,
                        last_name_lecturer,
                    )
                    raise LecturerNotFoundException(
                        f"Gutachter {first_name_lecturer} {last_name_lecturer} nicht gefunden."
                    )

                # jetzt alle Daten zusammen also Assignment erstellen
                assignment = self.create_assignment(
                    student.student_id,
                    lecturer.lecturer_id,
                    assignment_type,
                    topic,
                    grade,
                    assignment_date,
                    assignment_time,
                )
                imported_assignments.append(assignment)
                self.logger.debug(
                    "Assignment %s für Student Id %s importiert.",
                    assignment_type,
                    student.student_id,
                )

            return imported_assignments
        except Exception as e:
            self.logger.error("Error occurred: %s", e)
            raise e

    def generate_assignment_list(self, assignment_type, course_id):
        # generiert eine Liste mit allen Studenten eines Kurses um eine Übersicht
        # über die Projekt-/Bachelorarbeiten zu erstellen, welche später eingelesen werden kann

        csv_data = "Typ, Nachname Student, Vorname Student, Thema, Nachname Gutachter, Vorname Gutachter, Note, Datum, Uhrzeit\n"

        students_list = self.read_all_students_by_course_id(course_id)

        for student in students_list:
            if student.enrolled:
                csv_data += f"{assignment_type}, {student.last_name}, {student.first_name}, , , , ,  , \n"

        return csv_data

    def generate_assignment_list_excel(self, assignment_type, course_id):
        # generiert eine Excel-Datei mit allen Studenten eines Kurses um eine Übersicht
        # über die Projekt-/Bachelorarbeiten zu erstellen, welche später eingelesen werden kann
        wb = Workbook()
        ws = wb.active
        titles = "Typ, Nachname Student, Vorname Student, Thema, Nachname Gutachter, Vorname Gutachter, Note, Datum, Uhrzeit".split(
            ", "
        )

        course = self.read_course_by_id(course_id)
        ws.title = course.course_name
        students_list = self.read_all_students_by_course_id(course_id)

        # Schreibe Titel in die erste Zeile
        for col, title in enumerate(titles, start=1):
            ws.cell(row=1, column=col).value = title

        # Schreibe Studentendaten
        row_idx = 2
        for student in students_list:
            if student.enrolled:
                # Assignment ermitteln
                assignment = self.read_assignment_by_student_id_and_type(
                    student.student_id, assignment_type
                )
                if assignment:
                    if assignment.date:
                        assignment_date = datetime.strptime(assignment.date, "%Y-%m-%d %H:%M:%S").strftime("%d.%m.%Y")
                    else:
                        assignment_date = " "
                    if assignment.time:
                        assignment_time = assignment.time
                    else:
                        assignment_time = " "
                    # Gutachter ermitteln
                    lecturer = self.read_lecturer_by_id(assignment.lecturer_id)
                    if not lecturer:
                        self.logger.error(
                            "Gutachter mit der ID %s nicht gefunden.",
                            assignment.lecturer_id,
                        )
                        raise LecturerNotFoundException(
                            f"Gutachter {assignment.lecturer_id} nicht gefunden."
                        )

                # Schreibe Werte in die entsprechende Zeile
                ws.cell(row=row_idx, column=1).value = assignment_type
                ws.cell(row=row_idx, column=2).value = student.last_name
                ws.cell(row=row_idx, column=3).value = student.first_name
                if assignment:
                    ws.cell(row=row_idx, column=4).value = assignment.topic
                    ws.cell(row=row_idx, column=5).value = lecturer.last_name
                    ws.cell(row=row_idx, column=6).value = lecturer.first_name
                    ws.cell(row=row_idx, column=8).value = assignment_date
                    ws.cell(row=row_idx, column=9).value = assignment_time
                row_idx += 1
        self.logger.info("Es wurden %s Assignments in die Liste geschrieben.", row_idx - 1)
        return wb

    # Mail Methoden
    def get_mail_text(self, mail_type):
        if mail_type == "Freitext":
            text = (
                "Sie können nun einen beliebigen Text verfassen.\n"
                + "Bitte beachten Sie, dass die versendete Mail nicht ihrem Mail-Client angezeigt wird.\n"
                + "Besser Sie verwenden den studentischen Verteiler stud-verteiler+Kurs@lehre.dhbw-stuttgart.de."
            )
        else:
            text = (
                "Guten Tag {Vorname_student} {Nachname_student},\n"
                + "in der Zwischenzeit konnten für die meisten Arbeiten passende Gutachter*innen gefunden werden. "
                + "{Vorname_gutachter} {Nachname_gutachter} wird die wissenschaftliche Betreuung Ihrer "
                + str(mail_type)
                + " übernehmen.\n"
                + "Bitte nehmen Sie über die in CC stehende E-Mail-Adresse Kontakt mit Ihrem Gutachter auf, um sich über die Ausarbeitung der Arbeit abzustimmen.\n\n"
                + "Viel Erfolg und freundliche Grüße\n"
                + "Jonas Offtermatt "
            )

        return text

    def send_mail_to_students(self, email_subject, email_text, course_id):
        # versendet eine Mail mit dem Inhalt von email_text an alle eingeschriebenen Studenten im Kurs
        mail_client = DHBWMail()
        # alle Studenten aus dem Kurs lesen
        students_list = self.read_all_students_by_course_id(course_id)

        # Login Email Server
        try:
            mail_client.login()
            self.logger.debug("Erfolgreicher Login in Mail-Server")
        except Exception as e:
            self.logger.error("Error occurred: %s", e)
            result = {"success": False, "error_message": e}
            return result

        result = {"success": True, "error_message": ""}
        email_counter = 0
        for student in students_list:
            if student.enrolled:
                self.logger.debug(
                    "Sende Mail mit Betreff %s an %s", email_subject, student.email
                )
                result = mail_client.send_email_without_login(
                    student.email, subject=email_subject, body=email_text
                )
                if result["success"] is False:
                    self.logger.warning(
                        "Fehler beim Mailversand mit Betreff %s an %s",
                        email_subject,
                        student.email,
                    )
                    self.logger.error(
                        "Fehler beim Mailverand: %s", result["error_message"]
                    )
                    return result

                email_counter += 1
                self.logger.debug(
                            "Mail mit Betreff %s an %s (Student %s %s) erfolgreich versendet.",
                            email_subject,
                            student.email,
                            student.first_name,
                            student.last_name
                        )

        mail_client.logout()
        self.logger.info("Es wurden %s Emails erfolgreich versandt.", email_counter)
        return result

    def send_mail_to_students_and_lecturers(
        self, assignment_type, email_subject, email_text, course_id
    ):
        # versendet zu jeder studentischen Arbeit aus dem Kurs und vom gewählten Typ eine
        # Mail mit email_text und email_subject an alle eingeschriebenen Studenten + den
        # entsprechenden Gutachter
        # Dabei werden im Text noch die Platzhalter ersetzt.
        mail_client = DHBWMail()
        # alle Studenten aus dem Kurs lesen
        students_list = self.read_all_students_by_course_id(course_id)

        # Login Email Server
        try:
            mail_client.login()
            self.logger.debug("Erfolgreicher Login in Mail-Server")
        except Exception as e:
            self.logger.error("Error occurred: %s", e)
            result = {"success": False, "error_message": e}
            return result

        result = {"success": True, "error_message": ""}
        # Schleife ueber alle Studenten im Kurs
        email_counter = 0
        for student in students_list:
            if student.enrolled:
                # Assignment ermitteln
                assignment = self.read_assignment_by_student_id_and_type(
                    student.student_id, assignment_type
                )
                if assignment:
                    # Gutachter ermitteln
                    lecturer = self.read_lecturer_by_id(assignment.lecturer_id)
                    if not lecturer:
                        self.logger.error(
                            "Gutachter mit der ID %s nicht gefunden.",
                            assignment.lecturer_id,
                        )
                        raise LecturerNotFoundException(
                            f"Gutachter {assignment.lecturer_id} nicht gefunden."
                        )
                    # Platzhalter ersetzen
                    # Variablen für die Ersetzung
                    variables = {
                        "Vorname_student": student.first_name,
                        "Nachname_student": student.last_name,
                        "Vorname_gutachter": lecturer.first_name,
                        "Nachname_gutachter": lecturer.last_name,
                    }

                    # Ersetze die Platzhalter in der Zeichenkette durch die Variablen
                    formatted_email_text = email_text.format(**variables)
                    self.logger.debug(
                        "Sende Mail mit Betreff %s an %s", email_subject, student.email
                    )
                    result = mail_client.send_email_without_login(
                        to_address=student.email,
                        cc_address=lecturer.email,
                        subject=email_subject,
                        body=formatted_email_text,
                    )

                    if result["success"] is False:
                        self.logger.warning(
                            "Fehler beim Mailversand mit Betreff %s an %s",
                            email_subject,
                            student.email,
                        )
                        self.logger.error(
                            "Fehler beim Mailverand: %s", result["error_message"]
                        )
                        return result
                    else:
                        email_counter += 1
                        self.logger.debug(
                            "Mail mit Betreff %s an %s (Student %s %s) erfolgreich versendet.",
                            email_subject,
                            student.email,
                            student.first_name,
                            student.last_name
                        )
        # Ende Schleife ueber alle Studenten

        mail_client.logout()
        self.logger.info("Es wurden %s Emails erfolgreich versandt.", email_counter)
        return result

    # PDF Generierung Methoden
    def generate_evaluation_pdfs(self, save_path, assignment_type, course_id):
        # erzeugt für alle Assignments zum gewählten Type des Kurses ein Begutachtungsformular im Order save_path

        # Kursdaten ermitteln
        course = self.read_course_by_id(course_id)

        # alle Studenten aus dem Kurs lesen
        students_list = self.read_all_students_by_course_id(course_id)

        # Schleife ueber alle Studenten im Kurs
        generated_pdf_files = []
        for student in students_list:
            student_name = student.first_name + " " + student.last_name
            if student.enrolled:
                # Assignment ermitteln
                assignment = self.read_assignment_by_student_id_and_type(
                    student.student_id, assignment_type
                )
                if assignment:
                    if assignment.date:
                        assignment_date = datetime.strptime(assignment.date, "%Y-%m-%d %H:%M:%S").strftime("%d.%m.%Y")
                    else:
                        assignment_date = "  .  .        "
                    if assignment.time:
                        assignment_time = assignment.time
                    else:
                        assignment_time = "  :    -   :   "
                    # Gutachter ermitteln
                    lecturer = self.read_lecturer_by_id(assignment.lecturer_id)
                    lecturer_names = [
                        lecturer.first_name + " " + lecturer.last_name,
                    ]
                    if not lecturer:
                        self.logger.error(
                            "Gutachter mit der ID %s nicht gefunden.",
                            assignment.lecturer_id,
                        )
                        raise LecturerNotFoundException(
                            f"Gutachter {assignment.lecturer_id} nicht gefunden."
                        )

                    pdf_file = PresentationEvaluationPDF(
                        save_path,
                        assignment_type,
                        course.course_name,
                        assignment.topic,
                        lecturer_names,
                        student_name,
                        assignment_date,
                        assignment_time,
                    )
                    try:
                        pdf_file_path = pdf_file.create_evaluation()
                        self.logger.debug(
                            "Evaluations-PDF mit Thema %sfür Student %s erzeugt.",
                            assignment.topic,
                            student_name,
                        )
                        generated_pdf_files.append(pdf_file_path)
                    except Exception as e:
                        self.logger.error("Error occured: %s", e)
                        raise e
        self.logger.info(
            "Es wurden %s Evaluationsbögen erzeugt.", len(generated_pdf_files)
        )
        return generated_pdf_files
