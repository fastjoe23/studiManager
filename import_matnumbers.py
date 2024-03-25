# kleines Skript um vorhandene Studenten um Matrikelnummer anzureichern
from openpyxl import load_workbook
from studi_manager_controller import StudentManagerController
from studi_manager_model import Model

myModel = Model()
myController = StudentManagerController(myModel)


FILE_PATH = ""
workbook = load_workbook(FILE_PATH)

ws = workbook.active

for row in ws.iter_rows(min_row=2, values_only=True):
    matnumber = row[1]
    firstname = row[2]
    lastname = row[3]
    email = row[4]
    print("Lese Student", firstname, lastname, matnumber)

    # student aus Datenbank holen
    student = myController.read_student_by_name(lastname, firstname)

    if student:
        student_update = myController.update_student(
            student.student_id,
            student.person_id,
            student.last_name,
            student.first_name,
            student.email,
            student.company,
            matnumber,
            student.enrolled
        )
        print("Student aktualisiert: ", firstname, lastname)
    else:
        print("Student nicht gefunden: ", firstname, lastname)
        
